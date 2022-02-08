import requests
import json
import pandas as pd
import argparse
import time
import xlsxwriter
from openpyxl import load_workbook

# TODO How can I make this respect capital letters too?

parser = argparse.ArgumentParser()
parser.add_argument('--apikey', '-A', type=str, required=True, help='argument takes in the API key from web portal')

args = parser.parse_args()

def pp_json(json_thing, sort=False, indents=4):
    if type(json_thing) is str:
        print(json.dumps(json.loads(json_thing), sort_keys=sort, indent=indents))
    else:
        print(json.dumps(json_thing, sort_keys=sort, indent=indents))
    return None

#First Step is to get the bearer token from the api key. Need to pass it in.
#Should we make it a function?

api_url = "https://api.perimeter81.com/api/v1/auth/authorize"

data = '{"grantType": "api_key", "apiKey": "' + args.apikey + '"}'

headers = {'Content-type': 'application/json', 'accept': 'application/json'}

response = requests.post(api_url, headers=headers, data=data)

auth = {}
auth = json.loads(response.text)

token = auth['data']['accessToken']
#print(token)

bearerToken = {"Authorization": "Bearer " + token}
#print(bearerToken)

#print(auth['data']['accessTokenExpire'])
expires_time = auth['data']['accessTokenExpire']

base_url = 'https://api.perimeter81.com/api/rest'

json_header = {}
json_header['Authorization'] = 'Bearer ' +	 token
json_header['Accept'] = 'application/json'
json_header['Content-Type'] = 'application/json'

#xls_file = args.file
#data = pd.read_excel (args.file)
#df = pd.DataFrame(data, columns= ['GroupName', 'Description'])
#print(df)
#exit()

def createGroups(base_url, headers):
	url = '/v1/groups'

	#url = '/v1/users?page=1&qType=partial'
	#url = '/v1/users/?page=1&limit=25&q=Ace&qType=full'
	#for name in df_dict['GroupName']:
	for ind in df.index:

		#print(df['GroupName'][ind], df['Description'][ind])
		group_name = df['GroupName'][ind]
		group_desc = df['Description'][ind]
		calldata = '{"name": "' + group_name + '", "description": "' + group_desc + '"}'

		#print(calldata)
		#print(base_url+url)
		#print("###################")
		#print(headers)
		#print("GOT HEADERS")
		print("Creating group " + group_name)
		response = requests.post(base_url+url, headers=headers, data=calldata)
		time.sleep(5)
		print(response.text)

def listGroups(base_url, headers):

	id_to_username = {}

	writer = pd.ExcelWriter('users.xlsx', engine='xlsxwriter')
	writer.save()
	wb = load_workbook("users.xlsx")

	# Select First Worksheet
	ws = wb.worksheets[0]
	ws['A1'].value = 'ID'
	ws['B1'].value = 'Email'
	ws['C1'].value = 'Role'
	ws['D1'].value = 'First Name'
	ws['E1'].value = 'Last Name'
	ws['F1'].value = 'Groups'
	ws['F1'].value = 'Last Login'

	url = '/v1/users'
	response = requests.get(base_url + url, headers=headers)

	user_json = json.loads(response.text)
	#pp_json(user_json)

	for page in range(1, user_json['totalPage']+1):
		url = '/v1/users?page=' + str(page)
		response = requests.get(base_url + url, headers=headers)
		user_json = json.loads(response.text)
		pp_json(user_json)
		for user in user_json['data']:
			if user['terminated'] == False:
				#print(each['username'])
				#print(each['id'])
				id_to_username[user['id']] = user['username']

				if "lastVPNSuccessLoginAt" in user:
					new_row_data = [[user['id'], user['username'], user['role'], user['firstName'], user['lastName'],
									 user['lastVPNSuccessLoginAt']]]
				else:
					new_row_data = [[user['id'], user['username'], user['role'], user['firstName'], user['lastName']]]

				# Append 2 new Rows - Columns A - D
				for row_data in new_row_data:
					# Append Row Values
					ws.append(row_data)


				#print(user['username'])
				#print(user['tenantId'])
				#print(user)
	ws.title = user['tenantId']

	#pp_json(id_to_username)

	#exit()

	wb.create_sheet('Groups')

	wb.save("users.xlsx")

	#print(wb.sheetnames)

	ws = wb['Groups']
	ws['A1'].value = 'groupName'
	ws['B1'].value = 'Users'

	url = '/v1/groups'
	response = requests.get(base_url + url, headers=headers)
	groups_json = json.loads(response.text)
	#pp_json(groups_json['data'])
	#This will list all the groups and their members
	#print(type(groups_json))
	#print(type(groups_json['data']))
	for group in groups_json['data']:
		plain_user_list_string = ""
		#print(group['name'])
		for user in group['users']:
			#print("GOT HERE")
			#print(user)
			if user in id_to_username.keys():
				plain_user_list_string += (id_to_username[user] + "\n")
				first_name_last_name = id_to_username[user]
		#print(group['name'])
		#print("      " + plain_user_list_string)

		#print("     " + plain_user_list_string)
		new_row_data = [[group['name'], plain_user_list_string.rstrip("\n")]]

		#print(new_row_data)
		for row_data in new_row_data:
			# Append Row Values
			ws.append(row_data)

	wb.save("users.xlsx")

def backupUsers(base_url, headers):

	# Close the Pandas Excel writer and output the Excel file.

	#url = '/v1/users?page=1&limit=50'
	url = '/v1/users'
	response = requests.get(base_url + url, headers=headers)

	json_string = json.loads(response.text)
	#pp_json(json_string)
	#exit()
	#pp_json(json_string)
	#pp_json(json_string['data'][0]['idProviders']['azureAD']['groups'])
	#pp_json(json_string['totalPage'])
	writer = pd.ExcelWriter('demo.xlsx', engine='xlsxwriter')
	writer.save()
	wb = load_workbook("demo.xlsx")

	counter = 0
	for page in range(1, json_string['totalPage']+1):
		#url = '/v1/users?page=' + str(page) + '&limit=50'
		url = '/v1/users?page=' + str(page)
		#print(url)
		response = requests.get(base_url + url, headers=headers)
		json_string_page = json.loads(response.text)
		#pp_json(json_string_page)
		for user in json_string_page['data']:
			if not user['terminated']:
				print(user['username'])
				pp_json(user)
				#exit()
				counter += 1
				#print(user['username'])
				# new dataframe with same columns
				list_of_idps = []
				user_groups = {}
				#print(user['username'] + " is terminated:" + str(user['terminated']))
				for idp in user['idProviders']:
					if not idp == 'database':
						print(idp)
						if "groups" in user['idProviders'][idp]:
							print(user['idProviders'][idp]['groups'])
							user_groups[idp] = user['idProviders'][idp]['groups']

						#if user['idProviders'][each]['groups']:
						#	print(user['idProviders'][each]['groups'])
					#print(user['username'], user['role'], user['firstName'], user['lastName'])
					#pp_json(user_groups)

				# Select First Worksheet
				ws = wb.worksheets[0]
				ws.title = user['tenantId']
				ws['A1'].value = 'ID'
				ws['B1'].value = 'Email'
				ws['C1'].value = 'Role'
				ws['D1'].value = 'First Name'
				ws['E1'].value = 'Last Name'
				ws['K1'].value = 'Last Login'

				if "lastVPNSuccessLoginAt" in user:
					new_row_data = [[user['id'], user['username'], user['role'], user['firstName'], user['lastName'], user['lastVPNSuccessLoginAt']]]
				else:
					new_row_data = [[user['id'], user['username'], user['role'], user['firstName'], user['lastName']]]

				# Append 2 new Rows - Columns A - D
				for row_data in new_row_data:
					# Append Row Values
					ws.append(row_data)


	print(counter)
	wb.create_sheet('Groups')

	wb.save("users.xlsx")


	# 		# if checkKey(user['idProviders'], 'azureAD'):
	# 		# 	for group in user['idProviders']['azureAD']['groups']:
	# 		# 		#if "P81" in group:
	# 		# 		print("       " + group)
	# 	#pp_json(json_string_page['data'])

def listUsers(base_url, headers):
	#url = '/v1/users?page=1&limit=50'
	url = '/v1/users'
	response = requests.get(base_url + url, headers=headers)
	json_string = json.loads(response.text)
	pp_json(json_string)
	exit()
	#pp_json(json_string['data'][0]['idProviders']['azureAD']['groups'])
	#pp_json(json_string['totalPage'])
	for page in range(1, json_string['totalPage']+1):
		#url = '/v1/users?page=' + str(page) + '&limit=50'
		url = '/v1/users?page=' + str(page)
		#print(url)
		response = requests.get(base_url + url, headers=headers)
		json_string_page = json.loads(response.text)
		for user in json_string_page['data']:
			print(user['username'])
			if checkKey(user['idProviders'], 'azureAD'):
				for group in user['idProviders']['azureAD']['groups']:
					#if "P81" in group:
					print("       " + group)
		#pp_json(json_string_page['data'])

#for i in pp_json(json_string['data'])

def checkKey(dict, key):
	if key in dict.keys():
		#print(" ")
		#print("Present, ", end=" ")
		#print("value =", dict[key])
		return True
		#pass
	else:
		#pass
		return False
		#print(" ")
		#print("Not present")


# def shuffle(base_url, headers):
# 	url = 'v1/groups/' + GROUP_ID + '/member/' + USER_ID
# 	response = requests.get(base_url + url, headers=headers)
#


#createGroups(base_url, json_header)
listGroups(base_url, json_header)

#backupUsers(base_url, json_header)



