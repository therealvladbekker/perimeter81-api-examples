import requests
import json
import pandas as pd
import argparse
import time
import xlsxwriter
from openpyxl import load_workbook


parser = argparse.ArgumentParser()
parser.add_argument('--apikey', '-A', type=str, required=True, help='argument takes in the API key from web portal')

args = parser.parse_args()

def pp_json(json_thing, sort=False, indents=4):
    if type(json_thing) is str:
        print(json.dumps(json.loads(json_thing), sort_keys=sort, indent=indents))
    else:
        print(json.dumps(json_thing, sort_keys=sort, indent=indents))
    return None

#First Step is to get the bearer token from the api key. Need to pass it in.
#Should we make it a function?

api_url = "https://api.perimeter81.com/api/v1/auth/authorize"

data = '{"grantType": "api_key", "apiKey": "' + args.apikey + '"}'

#print(data)
#exit()

headers = {'Content-type': 'application/json', 'accept': 'application/json'}

response = requests.post(api_url, headers=headers, data=data)

auth = {}
auth = json.loads(response.text)
#print(auth)
#print(type(auth))

token = auth['data']['accessToken']
#print(token)

bearerToken = {"Authorization": "Bearer " + token}
#print(bearerToken)

#print(auth['data']['accessTokenExpire'])
expires_time = auth['data']['accessTokenExpire']


#datetime_time = datetime.datetime.fromtimestamp(expires_time)
#print(datetime_time)

base_url = 'https://api.perimeter81.com/api/rest'

json_header = {}
json_header['Authorization'] = 'Bearer ' +	 token
json_header['Accept'] = 'application/json'
json_header['Content-Type'] = 'application/json'
#json_header['Authorization'] = 'accessToken ' + token
#json_header['auth'] = 'Bearer ' + token


#xls_file = args.file
#data = pd.read_excel (args.file)
#df = pd.DataFrame(data, columns= ['GroupName', 'Description'])
#print(df)
#exit()

def createGroups(base_url, headers):
	url = '/v1/groups'

	#url = '/v1/users?page=1&qType=partial'
	#url = '/v1/users/?page=1&limit=25&q=Ace&qType=full'
	#for name in df_dict['GroupName']:
	for ind in df.index:

		#print(df['GroupName'][ind], df['Description'][ind])
		group_name = df['GroupName'][ind]
		group_desc = df['Description'][ind]
		calldata = '{"name": "' + group_name + '", "description": "' + group_desc + '"}'

		#print(calldata)
		#print(base_url+url)
		#print("###################")
		#print(headers)
		#print("GOT HEADERS")
		print("Creating group " + group_name)
		response = requests.post(base_url+url, headers=headers, data=calldata)
		time.sleep(5)
		print(response.text)

def listGroups(base_url, headers):
	url = '/v1/groups'
	response = requests.get(base_url + url, headers=headers)
	pp_json(response.text)

def backupUsers(base_url, headers):

	# Close the Pandas Excel writer and output the Excel file.

	#url = '/v1/users?page=1&limit=50'
	url = '/v1/users'
	response = requests.get(base_url + url, headers=headers)
	json_string = json.loads(response.text)
	#pp_json(json_string)
	#exit()
	#pp_json(json_string)
	#pp_json(json_string['data'][0]['idProviders']['azureAD']['groups'])
	#pp_json(json_string['totalPage'])
	writer = pd.ExcelWriter('demo.xlsx', engine='xlsxwriter')
	writer.save()
	wb = load_workbook("demo.xlsx")

	counter = 0
	for page in range(1, json_string['totalPage']+1):
		#url = '/v1/users?page=' + str(page) + '&limit=50'
		url = '/v1/users?page=' + str(page)
		#print(url)
		response = requests.get(base_url + url, headers=headers)
		json_string_page = json.loads(response.text)
		#pp_json(json_string_page)
		for user in json_string_page['data']:
			if not user['terminated']:
				print(user['username'])
				pp_json(user)
				#exit()
				counter += 1
				#print(user['username'])
				# new dataframe with same columns
				list_of_idps = []
				user_groups = {}
				#print(user['username'] + " is terminated:" + str(user['terminated']))
				for idp in user['idProviders']:
					if not idp == 'database':
						print(idp)
						if "groups" in user['idProviders'][idp]:
							print(user['idProviders'][idp]['groups'])
							user_groups[idp] = user['idProviders'][idp]['groups']

						#if user['idProviders'][each]['groups']:
						#	print(user['idProviders'][each]['groups'])
					#print(user['username'], user['role'], user['firstName'], user['lastName'])
					#pp_json(user_groups)

				# Select First Worksheet
				ws = wb.worksheets[0]
				ws.title = user['tenantId']
				ws['A1'].value = 'ID'
				ws['B1'].value = 'Email'
				ws['C1'].value = 'Role'
				ws['D1'].value = 'First Name'
				ws['E1'].value = 'Last Name'
				ws['F1'].value = 'Groups'
				ws['G1'].value = 'Last Login'

				if "lastVPNSuccessLoginAt" in user:
					new_row_data = [[ user['id'], user['username'], user['role'], user['firstName'], user['lastName'], str(user_groups), user['lastVPNSuccessLoginAt']]]
				else:
					new_row_data = [[ user['id'], user['username'], user['role'], user['firstName'], user['lastName'], str(user_groups)]]

				# Append 2 new Rows - Columns A - D
				for row_data in new_row_data:
					# Append Row Values
					ws.append(row_data)

	wb.save("demo.xlsx")

	print(counter)

	# 		# if checkKey(user['idProviders'], 'azureAD'):
	# 		# 	for group in user['idProviders']['azureAD']['groups']:
	# 		# 		#if "P81" in group:
	# 		# 		print("       " + group)
	# 	#pp_json(json_string_page['data'])

def listUsers(base_url, headers):
	#url = '/v1/users?page=1&limit=50'
	url = '/v1/users'
	response = requests.get(base_url + url, headers=headers)
	json_string = json.loads(response.text)
	pp_json(json_string)
	exit()
	#pp_json(json_string['data'][0]['idProviders']['azureAD']['groups'])
	#pp_json(json_string['totalPage'])
	for page in range(1, json_string['totalPage']+1):
		#url = '/v1/users?page=' + str(page) + '&limit=50'
		url = '/v1/users?page=' + str(page)
		#print(url)
		response = requests.get(base_url + url, headers=headers)
		json_string_page = json.loads(response.text)
		for user in json_string_page['data']:
			print(user['username'])
			if checkKey(user['idProviders'], 'azureAD'):
				for group in user['idProviders']['azureAD']['groups']:
					#if "P81" in group:
					print("       " + group)
		#pp_json(json_string_page['data'])

#for i in pp_json(json_string['data'])

def checkKey(dict, key):
	if key in dict.keys():
		#print(" ")
		#print("Present, ", end=" ")
		#print("value =", dict[key])
		return True
		#pass
	else:
		#pass
		return False
		#print(" ")
		#print("Not present")


# def shuffle(base_url, headers):
# 	url = 'v1/groups/' + GROUP_ID + '/member/' + USER_ID
# 	response = requests.get(base_url + url, headers=headers)
#


#createGroups(base_url, json_header)
#listGroups(base_url, json_header)

backupUsers(base_url, json_header)



